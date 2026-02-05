"""
FRED Data Fetcher

Fetches economic data from the Federal Reserve Economic Data (FRED) API.
Supports both initial bulk download and incremental updates.

API Documentation: https://fred.stlouisfed.org/docs/api/fred/
"""

import requests
import json
import time
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import logging
from dotenv import load_dotenv
import os

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class FREDFetcher:
    """Handles fetching data from FRED API"""
    
    # FRED series configuration - matches Excel file series
    FRED_SERIES = {
        'MORTGAGE30US': {
            'name': '30-Year Fixed Rate Mortgage Average',
            'frequency': 'weekly',
            'units': 'percent',
            'sheet': 'Weekly, Ending Thursday'
        },
        'CPIAUCSL': {
            'name': 'Consumer Price Index for All Urban Consumers',
            'frequency': 'monthly',
            'units': 'index',
            'sheet': 'Monthly'
        },
        'RHORUSQ156N': {
            'name': 'Homeownership Rate in the United States',
            'frequency': 'quarterly',
            'units': 'percent',
            'sheet': 'Quarterly'
        },
        'HPIPONM226S_PCH': {
            'name': 'Purchase Only House Price Index',
            'frequency': 'monthly',
            'units': 'percent_change',
            'sheet': 'Monthly'
        },
        'GDPC1CTM': {
            'name': 'GDPC1CTM',
            'frequency': 'annual',
            'units': 'unknown',
            'sheet': 'Annual'
        }
    }
    
    BASE_URL = 'https://api.stlouisfed.org/fred'
    RATE_LIMIT_DELAY = 0.5  # 500ms between requests (120 req/min limit)
    
    def __init__(self, api_key: str, raw_dir: str):
        """
        Initialize FRED fetcher
        
        Args:
            api_key: FRED API key
            raw_dir: Directory to save raw JSON data
        """
        self.api_key = api_key
        self.raw_dir = Path(raw_dir)
        self.raw_dir.mkdir(parents=True, exist_ok=True)
        
        # Track last request time for rate limiting
        self.last_request_time = 0
        
    def _rate_limit(self):
        """Enforce rate limiting between API requests"""
        elapsed = time.time() - self.last_request_time
        if elapsed < self.RATE_LIMIT_DELAY:
            time.sleep(self.RATE_LIMIT_DELAY - elapsed)
        self.last_request_time = time.time()
    
    def _make_request(self, endpoint: str, params: Dict) -> Dict:
        """
        Make API request with rate limiting and error handling
        
        Args:
            endpoint: API endpoint (e.g., 'series/observations')
            params: Query parameters
            
        Returns:
            JSON response as dictionary
        """
        self._rate_limit()
        
        # Add API key to params
        params['api_key'] = self.api_key
        params['file_type'] = 'json'
        
        url = f"{self.BASE_URL}/{endpoint}"
        
        try:
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"API request failed: {e}")
            raise
    
    def get_series_info(self, series_id: str) -> Dict:
        """
        Get metadata for a FRED series
        
        Args:
            series_id: FRED series ID
            
        Returns:
            Series metadata
        """
        logger.info(f"Fetching metadata for {series_id}...")
        return self._make_request('series', {'series_id': series_id})
    
    def get_observations(
        self, 
        series_id: str, 
        start_date: Optional[str] = None,
        end_date: Optional[str] = None
    ) -> List[Dict]:
        """
        Get observations for a FRED series
        
        Args:
            series_id: FRED series ID
            start_date: Start date (YYYY-MM-DD), defaults to earliest available
            end_date: End date (YYYY-MM-DD), defaults to latest available
            
        Returns:
            List of observations
        """
        params = {'series_id': series_id}
        
        if start_date:
            params['observation_start'] = start_date
        if end_date:
            params['observation_end'] = end_date
        
        logger.info(f"Fetching observations for {series_id}...")
        if start_date:
            logger.info(f"  Date range: {start_date} to {end_date or 'latest'}")
        
        response = self._make_request('series/observations', params)
        observations = response.get('observations', [])
        
        logger.info(f"  Retrieved {len(observations)} observations")
        
        return observations
    
    def get_last_update_date(self, series_id: str) -> Optional[str]:
        """
        Get the date of the last fetched observation for a series
        
        Args:
            series_id: FRED series ID
            
        Returns:
            Last observation date (YYYY-MM-DD) or None if no data exists
        """
        metadata_file = self.raw_dir / f"{series_id}_metadata.json"
        
        if not metadata_file.exists():
            return None
        
        try:
            with open(metadata_file, 'r') as f:
                metadata = json.load(f)
                return metadata.get('last_observation_date')
        except Exception as e:
            logger.warning(f"Could not read metadata for {series_id}: {e}")
            return None
    
    def save_series_data(self, series_id: str, observations: List[Dict], metadata: Dict):
        """
        Save series data and metadata to JSON files
        
        Args:
            series_id: FRED series ID
            observations: List of observations
            metadata: Series metadata
        """
        # Save observations
        obs_file = self.raw_dir / f"{series_id}_observations.json"
        with open(obs_file, 'w') as f:
            json.dump(observations, f, indent=2)
        
        # Save metadata with last update info
        metadata_file = self.raw_dir / f"{series_id}_metadata.json"
        
        # Find last observation date
        valid_obs = [obs for obs in observations if obs['value'] != '.']
        last_date = max([obs['date'] for obs in valid_obs]) if valid_obs else None
        
        metadata_to_save = {
            'series_id': series_id,
            'name': self.FRED_SERIES[series_id]['name'],
            'frequency': self.FRED_SERIES[series_id]['frequency'],
            'units': self.FRED_SERIES[series_id]['units'],
            'description': self.FRED_SERIES[series_id]['description'],
            'last_observation_date': last_date,
            'observation_count': len(observations),
            'fetched_at': datetime.now().isoformat(),
            'fred_metadata': metadata
        }
        
        with open(metadata_file, 'w') as f:
            json.dump(metadata_to_save, f, indent=2)
        
        logger.info(f"  Saved to {obs_file}")
    
    def fetch_series(self, series_id: str, incremental: bool = True):
        """
        Fetch data for a single series
        
        Args:
            series_id: FRED series ID
            incremental: If True, only fetch data since last update
        """
        logger.info(f"\n{'='*60}")
        logger.info(f"Fetching {series_id}: {self.FRED_SERIES[series_id]['name']}")
        logger.info(f"{'='*60}")
        
        # Get series metadata
        series_info = self.get_series_info(series_id)
        
        # Determine start date for incremental updates
        start_date = None
        if incremental:
            last_date = self.get_last_update_date(series_id)
            if last_date:
                # Fetch from day after last observation
                last_dt = datetime.strptime(last_date, '%Y-%m-%d')
                start_dt = last_dt + timedelta(days=1)
                start_date = start_dt.strftime('%Y-%m-%d')
                logger.info(f"Incremental update from {start_date}")
        
        # Fetch observations
        observations = self.get_observations(series_id, start_date=start_date)
        
        if not observations:
            logger.info("  No new observations")
            return
        
        # If incremental, merge with existing data
        if incremental and start_date:
            existing_file = self.raw_dir / f"{series_id}_observations.json"
            if existing_file.exists():
                with open(existing_file, 'r') as f:
                    existing_obs = json.load(f)
                
                # Combine and deduplicate
                all_obs = existing_obs + observations
                seen_dates = set()
                unique_obs = []
                for obs in all_obs:
                    if obs['date'] not in seen_dates:
                        unique_obs.append(obs)
                        seen_dates.add(obs['date'])
                
                observations = sorted(unique_obs, key=lambda x: x['date'])
                logger.info(f"  Merged with existing data: {len(observations)} total observations")
        
        # Save data
        self.save_series_data(series_id, observations, series_info)
    
    def fetch_all(self, incremental: bool = True):
        """
        Fetch all configured FRED series
        
        Args:
            incremental: If True, only fetch new data since last update
        """
        logger.info("Starting FRED data fetch...")
        logger.info(f"Mode: {'Incremental' if incremental else 'Full download'}")
        logger.info(f"Series to fetch: {len(self.FRED_SERIES)}")
        
        for series_id in self.FRED_SERIES.keys():
            try:
                self.fetch_series(series_id, incremental=incremental)
            except Exception as e:
                logger.error(f"Error fetching {series_id}: {e}")
                # Continue with other series
        
        logger.info("\n✅ FRED data fetch complete!")
    
    def update_excel_and_upload(self, excel_file: str):
        """
        Fetch latest FRED data, update Excel file, and trigger upload to BigQuery
        
        Args:
            excel_file: Path to Excel file to update
        """
        import pandas as pd
        import subprocess
        from openpyxl import load_workbook
        
        logger.info("\n" + "="*60)
        logger.info("FRED Data Update & Upload Pipeline")
        logger.info("="*60)
        
        # Step 1: Fetch latest data from FRED API
        logger.info("\n📥 Step 1: Fetching latest data from FRED API...")
        self.fetch_all(incremental=True)
        
        # Step 2: Update Excel file with new data
        logger.info("\n📝 Step 2: Updating Excel file...")
        excel_path = Path(excel_file)
        
        if not excel_path.exists():
            logger.error(f"Excel file not found: {excel_file}")
            return
        
        # Load workbook
        wb = load_workbook(excel_path)
        
        # Update each sheet with new data from API
        for series_id, config in self.FRED_SERIES.items():
            obs_file = self.raw_dir / f"{series_id}_observations.json"
            
            if not obs_file.exists():
                logger.warning(f"No data file for {series_id}, skipping...")
                continue
            
            # Load observations from JSON
            with open(obs_file, 'r') as f:
                observations = json.load(f)
            
            # Convert to DataFrame
            df = pd.DataFrame(observations)
            df = df[df['value'] != '.']  # Remove missing values
            df['value'] = pd.to_numeric(df['value'])
            df['date'] = pd.to_datetime(df['date'])
            
            # Get the sheet name
            sheet_name = config['sheet']
            
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in Excel, skipping {series_id}")
                continue
            
            ws = wb[sheet_name]
            
            # Find the column for this series (header row is row 1)
            series_col = None
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value == series_id:
                    series_col = col_idx
                    break
            
            if series_col is None:
                logger.warning(f"Column for {series_id} not found in sheet '{sheet_name}'")
                continue
            
            # Update the column with new data
            # Find existing dates and append new ones
            existing_dates = set()
            for row_idx in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row_idx, column=1)
                if date_cell.value:
                    existing_dates.add(pd.to_datetime(date_cell.value).date())
            
            # Add new observations
            new_count = 0
            for _, row in df.iterrows():
                obs_date = row['date'].date()
                if obs_date not in existing_dates:
                    # Find next empty row
                    next_row = ws.max_row + 1
                    ws.cell(row=next_row, column=1, value=row['date'])
                    ws.cell(row=next_row, column=series_col, value=row['value'])
                    new_count += 1
            
            logger.info(f"  {series_id}: Added {new_count} new observations")
        
        # Save updated Excel file
        wb.save(excel_path)
        logger.info(f"\n✅ Excel file updated: {excel_path}")
        
        # Step 3: Run preprocessing
        logger.info("\n🔄 Step 3: Running preprocessing...")
        preprocess_script = Path(__file__).parent / 'preprocess_fred.py'
        result = subprocess.run(['python', str(preprocess_script)], capture_output=True, text=True)
        
        if result.returncode != 0:
            logger.error(f"Preprocessing failed: {result.stderr}")
            return
        
        logger.info("✅ Preprocessing complete")
        
        # Step 4: Upload to BigQuery
        logger.info("\n☁️  Step 4: Uploading to BigQuery...")
        upload_script = Path(__file__).parent / 'upload_fred_to_bigquery.py'
        result = subprocess.run(['python', str(upload_script)], capture_output=True, text=True)
        
        if result.returncode != 0:
            logger.error(f"Upload failed: {result.stderr}")
            return
        
        logger.info("✅ Upload complete")
        
        logger.info("\n" + "="*60)
        logger.info("🎉 FRED data update pipeline complete!")
        logger.info("="*60)


def main():
    """Main execution function"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Fetch FRED data and update BigQuery')
    parser.add_argument('--mode', choices=['fetch-only', 'full-update'], default='full-update',
                       help='fetch-only: Just download to JSON. full-update: Update Excel and upload to BigQuery')
    args = parser.parse_args()
    
    # Load API key from environment
    api_key = os.getenv('FRED_API_KEY')
    
    if not api_key:
        logger.error("FRED_API_KEY not found in environment variables")
        logger.error("Please set FRED_API_KEY in your .env file")
        return
    
    # Define paths
    raw_dir = Path(__file__).parent / 'fred_raw'
    excel_file = Path(__file__).parent / 'historic_fred' / 'fred_macro_data.xlsx'
    
    # Create fetcher
    fetcher = FREDFetcher(
        api_key=api_key,
        raw_dir=str(raw_dir)
    )
    
    if args.mode == 'fetch-only':
        # Just fetch and save to JSON
        fetcher.fetch_all(incremental=True)
    else:
        # Full pipeline: fetch, update Excel, preprocess, and upload
        fetcher.update_excel_and_upload(str(excel_file))


if __name__ == '__main__':
    main()
